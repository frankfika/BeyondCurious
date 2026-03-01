#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成尽调财务分析 Excel 文件"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = "/Users/fangchen/Desktop/测试数据/模拟尽调资料包"
OUTPUT_DIR = os.path.join(BASE_DIR, "10 财务分析")

# 样式定义
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True)
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True)
NORMAL_FONT = Font(name='微软雅黑', size=10)
CURRENCY_FONT = Font(name='微软雅黑', size=10)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
ALT_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def create_financial_statements():
    """创建财务报表（三表合一）"""
    wb = Workbook()

    # ===== 资产负债表 =====
    ws1 = wb.active
    ws1.title = "资产负债表"

    ws1['A1'] = "资产负债表"
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells('A1:F1')
    ws1['A2'] = "编制单位：杭州云帆智能科技有限公司                    2023年12月31日                    单位：万元"
    ws1.merge_cells('A2:F2')

    headers = ['资产', '期末余额', '期初余额', '负债和所有者权益', '期末余额', '期初余额']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # 资产数据
    assets = [
        ('流动资产：', '', '', '流动负债：', '', ''),
        ('  货币资金', 2856.50, 1923.80, '  短期借款', 1500.00, 1200.00),
        ('  交易性金融资产', 500.00, 300.00, '  应付票据', 320.00, 280.00),
        ('  应收票据', 186.00, 145.00, '  应付账款', 1020.00, 856.00),
        ('  应收账款', 3629.35, 2305.80, '  预收款项', 456.00, 380.00),
        ('  预付款项', 125.00, 98.00, '  应付职工薪酬', 186.50, 156.80),
        ('  其他应收款', 85.00, 62.00, '  应交税费', 245.00, 198.00),
        ('  存货', 425.00, 356.00, '  其他应付款', 98.00, 75.00),
        ('  其他流动资产', 45.00, 32.00, '  一年内到期非流动负债', 200.00, 150.00),
        ('流动资产合计', 7851.85, 5222.60, '流动负债合计', 4025.50, 3295.80),
        ('', '', '', '', '', ''),
        ('非流动资产：', '', '', '非流动负债：', '', ''),
        ('  长期股权投资', 500.00, 500.00, '  长期借款', 800.00, 600.00),
        ('  固定资产', 1856.00, 1625.00, '  递延收益', 350.00, 280.00),
        ('  在建工程', 320.00, 180.00, '  其他非流动负债', 50.00, 40.00),
        ('  无形资产', 680.00, 520.00, '非流动负债合计', 1200.00, 920.00),
        ('  长期待摊费用', 125.00, 98.00, '', '', ''),
        ('  递延所得税资产', 85.00, 65.00, '负债合计', 5225.50, 4215.80),
        ('非流动资产合计', 3566.00, 2988.00, '', '', ''),
        ('', '', '', '所有者权益：', '', ''),
        ('', '', '', '  实收资本', 5000.00, 5000.00),
        ('', '', '', '  资本公积', 850.00, 850.00),
        ('', '', '', '  盈余公积', 156.00, 98.00),
        ('', '', '', '  未分配利润', 186.35, 46.80),
        ('', '', '', '所有者权益合计', 6192.35, 5994.80),
        ('资产总计', 11417.85, 8210.60, '负债和所有者权益总计', 11417.85, 8210.60),
    ]

    for row_idx, row_data in enumerate(assets, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [2, 3, 5, 6] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if '合计' in str(row_data[0]) or '总计' in str(row_data[0]):
                cell.font = HEADER_FONT
                if col_idx in [1, 4]:
                    cell.fill = TOTAL_FILL

    set_column_width(ws1, [22, 14, 14, 22, 14, 14])

    # ===== 利润表 =====
    ws2 = wb.create_sheet("利润表")
    ws2['A1'] = "利润表"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:D1')
    ws2['A2'] = "编制单位：杭州云帆智能科技有限公司                    2023年度                    单位：万元"
    ws2.merge_cells('A2:D2')

    headers2 = ['项目', '本期金额', '上期金额', '增减比例']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    profit_data = [
        ('一、营业收入', 12568.50, 9856.20, '27.6%'),
        ('  减：营业成本', 6892.30, 5562.80, '23.9%'),
        ('  税金及附加', 125.68, 98.56, '27.6%'),
        ('  销售费用', 1125.50, 856.20, '31.5%'),
        ('  管理费用', 1456.80, 1125.60, '29.4%'),
        ('  研发费用', 1856.50, 1425.80, '30.2%'),
        ('  财务费用', 186.50, 125.60, '48.5%'),
        ('  其中：利息费用', 156.00, 105.00, '48.6%'),
        ('  加：其他收益', 256.00, 186.00, '37.6%'),
        ('  投资收益', 45.00, 32.00, '40.6%'),
        ('二、营业利润', 1228.22, 881.64, '39.3%'),
        ('  加：营业外收入', 85.00, 56.00, '51.8%'),
        ('  减：营业外支出', 25.00, 18.00, '38.9%'),
        ('三、利润总额', 1288.22, 919.64, '40.1%'),
        ('  减：所得税费用', 193.23, 137.95, '40.1%'),
        ('四、净利润', 1094.99, 781.69, '40.1%'),
        ('', '', '', ''),
        ('五、每股收益', '', '', ''),
        ('  基本每股收益（元/股）', 0.219, 0.156, '40.4%'),
        ('  稀释每股收益（元/股）', 0.219, 0.156, '40.4%'),
    ]

    for row_idx, row_data in enumerate(profit_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [2, 3] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if str(row_data[0]).startswith(('一、', '二、', '三、', '四、', '五、')):
                cell.font = HEADER_FONT
                cell.fill = TOTAL_FILL

    set_column_width(ws2, [28, 16, 16, 12])

    # ===== 现金流量表 =====
    ws3 = wb.create_sheet("现金流量表")
    ws3['A1'] = "现金流量表"
    ws3['A1'].font = TITLE_FONT
    ws3.merge_cells('A1:D1')
    ws3['A2'] = "编制单位：杭州云帆智能科技有限公司                    2023年度                    单位：万元"
    ws3.merge_cells('A2:D2')

    headers3 = ['项目', '本期金额', '上期金额', '增减']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    cashflow_data = [
        ('一、经营活动产生的现金流量', '', '', ''),
        ('  销售商品、提供劳务收到的现金', 11256.80, 8652.50, 2604.30),
        ('  收到的税费返还', 156.00, 98.00, 58.00),
        ('  收到其他与经营活动有关的现金', 325.00, 256.00, 69.00),
        ('  经营活动现金流入小计', 11737.80, 9006.50, 2731.30),
        ('  购买商品、接受劳务支付的现金', 6258.50, 4856.20, 1402.30),
        ('  支付给职工以及为职工支付的现金', 2856.80, 2256.50, 600.30),
        ('  支付的各项税费', 685.60, 526.80, 158.80),
        ('  支付其他与经营活动有关的现金', 856.50, 652.30, 204.20),
        ('  经营活动现金流出小计', 10657.40, 8291.80, 2365.60),
        ('  经营活动产生的现金流量净额', 1080.40, 714.70, 365.70),
        ('', '', '', ''),
        ('二、投资活动产生的现金流量', '', '', ''),
        ('  收回投资收到的现金', 200.00, 150.00, 50.00),
        ('  取得投资收益收到的现金', 45.00, 32.00, 13.00),
        ('  处置固定资产收到的现金', 25.00, 18.00, 7.00),
        ('  投资活动现金流入小计', 270.00, 200.00, 70.00),
        ('  购建固定资产支付的现金', 526.00, 425.00, 101.00),
        ('  投资支付的现金', 500.00, 300.00, 200.00),
        ('  投资活动现金流出小计', 1026.00, 725.00, 301.00),
        ('  投资活动产生的现金流量净额', -756.00, -525.00, -231.00),
        ('', '', '', ''),
        ('三、筹资活动产生的现金流量', '', '', ''),
        ('  吸收投资收到的现金', 0.00, 0.00, 0.00),
        ('  取得借款收到的现金', 800.00, 600.00, 200.00),
        ('  筹资活动现金流入小计', 800.00, 600.00, 200.00),
        ('  偿还债务支付的现金', 400.00, 350.00, 50.00),
        ('  分配股利、利润支付的现金', 125.00, 85.00, 40.00),
        ('  偿付利息支付的现金', 156.00, 105.00, 51.00),
        ('  筹资活动现金流出小计', 681.00, 540.00, 141.00),
        ('  筹资活动产生的现金流量净额', 119.00, 60.00, 59.00),
        ('', '', '', ''),
        ('四、汇率变动对现金的影响', 12.70, 8.50, 4.20),
        ('五、现金及现金等价物净增加额', 456.10, 258.20, 197.90),
        ('  加：期初现金及现金等价物余额', 1923.80, 1665.60, 258.20),
        ('六、期末现金及现金等价物余额', 2379.90, 1923.80, 456.10),
    ]

    for row_idx, row_data in enumerate(cashflow_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [2, 3, 4] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if str(row_data[0]).startswith(('一、', '二、', '三、', '四、', '五、', '六、')):
                cell.font = HEADER_FONT
                cell.fill = TOTAL_FILL

    set_column_width(ws3, [36, 14, 14, 12])

    wb.save(os.path.join(OUTPUT_DIR, "财务报表.xlsx"))
    print("✓ 财务报表.xlsx (资产负债表、利润表、现金流量表)")

def create_receivables():
    """创建应收账款明细表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "应收账款明细"

    ws['A1'] = "应收账款明细表"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:I1')
    ws['A2'] = "截至2023年12月31日                                    单位：万元"
    ws.merge_cells('A2:I2')

    headers = ['序号', '客户名称', '合同编号', '应收金额', '账龄', '预计回收日期', '信用等级', '坏账准备', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, '杭州市数据资源管理局', 'HT-GOV-2023-028', 425.00, '3-6个月', '2024-03', 'AAA', 4.25, '政府项目验收后付款'),
        (2, '浙商银行股份有限公司', 'HT-FIN-2023-015', 380.00, '1-3个月', '2024-02', 'AAA', 3.80, '年度合同尾款'),
        (3, '浙江省交通投资集团有限公司', 'HT-TRA-2023-032', 285.00, '3-6个月', '2024-03', 'AA+', 14.25, '项目验收后付款'),
        (4, '宁波银行股份有限公司', 'HT-FIN-2023-022', 210.00, '1-3个月', '2024-02', 'AAA', 2.10, '正常回款中'),
        (5, '浙江大学医学院附属第一医院', 'HT-MED-2023-018', 186.00, '6-12个月', '2024-06', 'AA', 18.60, '医院付款流程较长'),
        (6, '杭州海康威视数字技术股份有限公司', 'HT-MFG-2023-025', 168.00, '1-3个月', '2024-02', 'AAA', 1.68, '履行中合同分期'),
        (7, '深圳前海微众银行股份有限公司', 'HT-FIN-2023-035', 152.00, '1-3个月', '2024-02', 'AAA', 1.52, '正常回款中'),
        (8, '浙江联华华商集团有限公司', 'HT-RET-2023-019', 145.00, '1-3个月', '2024-03', 'AA+', 1.45, '履行中合同分期'),
        (9, '中国人寿保险浙江省分公司', 'HT-FIN-2023-038', 285.00, '1-3个月', '2024-03', 'AAA', 2.85, '新签合同首期'),
        (10, '正泰集团股份有限公司', 'HT-MFG-2023-016', 186.00, '3-6个月', '2024-03', 'AA+', 9.30, '质保金'),
        (11, '浙江省能源集团有限公司', 'HT-ENE-2023-021', 125.00, '3-6个月', '2024-04', 'AA+', 6.25, '分期付款第三期'),
        (12, '阿里健康信息技术（浙江）有限公司', 'HT-MED-2023-029', 98.00, '1-3个月', '2024-02', 'AAA', 0.98, '正常回款中'),
        (13, '浙江省农村信用社联合社', 'HT-FIN-2023-042', 85.00, '1-3个月', '2024-02', 'AA+', 0.85, '新客户首期款'),
        (14, '网易（杭州）网络有限公司', 'HT-NET-2023-027', 68.00, '1-3个月', '2024-02', 'AAA', 0.68, '正常回款中'),
        (15, '杭州万事利丝绸文化股份有限公司', 'HT-MFG-2023-024', 52.00, '1-3个月', '2024-03', 'AA', 0.52, '正常回款中'),
        (16, '其他客户（汇总）', '各合同', 1092.00, '各期', '-', '各级', 92.52, '80家客户合计'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [4, 8] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 合计行
    total_row = 5 + len(data)
    ws.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws.cell(row=total_row, column=4, value=3892.00).font = HEADER_FONT
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=8, value=262.65).font = HEADER_FONT
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    for col in range(1, 10):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws, [6, 28, 18, 12, 10, 14, 10, 12, 22])

    wb.save(os.path.join(OUTPUT_DIR, "应收账款明细表.xlsx"))
    print("✓ 应收账款明细表.xlsx")

def create_payables():
    """创建应付账款明细表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "应付账款明细"

    ws['A1'] = "应付账款明细表"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    ws['A2'] = "截至2023年12月31日                                    单位：万元"
    ws.merge_cells('A2:F2')

    headers = ['序号', '供应商名称', '应付金额', '账龄', '计划付款日期', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, '阿里云计算有限公司', 125.00, '1个月内', '2024-01', '12月云服务费'),
        (2, '英伟达半导体科技（上海）有限公司', 260.00, '1-3个月', '2024-02', 'GPU采购尾款'),
        (3, '华为技术有限公司', 98.00, '1-3个月', '2024-02', '服务器采购款'),
        (4, '亚马逊云科技（北京）有限公司', 68.00, '1个月内', '2024-01', '12月AWS费用'),
        (5, '杭州数据标注服务有限公司', 42.00, '1个月内', '2024-01', '12月标注费'),
        (6, '上海智云网络科技有限公司', 35.00, '1-3个月', '2024-02', '安全服务费'),
        (7, '杭州优办物业管理有限公司', 27.00, '1个月内', '2024-01', 'Q4物业费'),
        (8, '中通服供应链管理有限公司', 85.00, '1-3个月', '2024-02', '设备采购款'),
        (9, '北京神州数码有限公司', 56.00, '1-3个月', '2024-03', '软件授权费'),
        (10, '杭州网新恒天软件有限公司', 38.00, '1个月内', '2024-01', '外包开发费'),
        (11, '中国移动通信集团浙江有限公司', 28.00, '1个月内', '2024-01', '通信服务费'),
        (12, '其他供应商（汇总）', 158.00, '各期', '-', '30家供应商'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx == 3 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 合计行
    total_row = 5 + len(data)
    ws.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws.merge_cells(f'B{total_row}:B{total_row}')
    ws.cell(row=total_row, column=3, value=1020.00).font = HEADER_FONT
    ws.cell(row=total_row, column=3).number_format = '#,##0.00'
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws, [6, 32, 12, 10, 14, 20])

    wb.save(os.path.join(OUTPUT_DIR, "应付账款明细表.xlsx"))
    print("✓ 应付账款明细表.xlsx")

def create_related_party():
    """创建关联交易明细表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "关联交易明细"

    ws['A1'] = "关联交易明细表"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')
    ws['A2'] = "2023年度                                           单位：万元"
    ws.merge_cells('A2:G2')

    headers = ['序号', '日期', '关联方', '交易类型', '金额', '定价依据', '审批状态']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, '2023-01', '杭州云帆智能软件有限公司', '技术服务费', 35.00, '参照市场价格', '已批准'),
        (2, '2023-02', '杭州云帆智能软件有限公司', '技术服务费', 38.00, '参照市场价格', '已批准'),
        (3, '2023-03', '杭州云帆数据科技有限公司', '技术服务费', 15.00, '参照市场价格', '已批准'),
        (4, '2023-03', '杭州云帆智能软件有限公司', '管理费', 18.00, '成本加成法', '已批准'),
        (5, '2023-04', '杭州云图科技有限公司', '技术合作', 50.00, '协商定价', '已批准'),
        (6, '2023-05', '杭州云帆智能软件有限公司', '技术服务费', 42.00, '参照市场价格', '已批准'),
        (7, '2023-06', '杭州云帆数据科技有限公司', '技术服务费', 18.00, '参照市场价格', '已批准'),
        (8, '2023-06', '杭州云帆智能软件有限公司', '管理费', 18.00, '成本加成法', '已批准'),
        (9, '2023-07', '杭州云帆数据科技有限公司', '管理费', 14.00, '成本加成法', '已批准'),
        (10, '2023-08', '杭州云图科技有限公司', '技术合作', 45.00, '协商定价', '已批准'),
        (11, '2023-09', '杭州云帆智能软件有限公司', '技术服务费', 38.00, '参照市场价格', '已批准'),
        (12, '2023-09', '杭州云帆智能软件有限公司', '管理费', 18.00, '成本加成法', '已批准'),
        (13, '2023-10', '杭州云帆数据科技有限公司', '技术服务费', 22.00, '参照市场价格', '已批准'),
        (14, '2023-11', '杭州云图科技有限公司', '技术合作', 55.00, '协商定价', '已批准'),
        (15, '2023-11', '杭州云图科技有限公司', '办公场地分摊', 18.00, '按面积分摊', '已批准'),
        (16, '2023-12', '杭州云帆智能软件有限公司', '技术服务费', 40.00, '参照市场价格', '已批准'),
        (17, '2023-12', '杭州云帆数据科技有限公司', '技术服务费', 16.00, '参照市场价格', '已批准'),
        (18, '2023-12', '杭州云帆智能软件有限公司', '管理费', 18.00, '成本加成法', '已批准'),
        (19, '2023-12', '杭州云帆数据科技有限公司', '管理费', 14.00, '成本加成法', '已批准'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx == 5 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 合计行
    total_row = 5 + len(data)
    ws.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws.merge_cells(f'B{total_row}:D{total_row}')
    ws.cell(row=total_row, column=5, value=893.00).font = HEADER_FONT
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws, [6, 12, 26, 12, 10, 14, 10])

    wb.save(os.path.join(OUTPUT_DIR, "关联交易明细表.xlsx"))
    print("✓ 关联交易明细表.xlsx")

def create_revenue():
    """创建收入明细表"""
    wb = Workbook()

    # 按产品分类
    ws1 = wb.active
    ws1.title = "按产品分类"

    ws1['A1'] = "分产品收入统计表"
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells('A1:F1')
    ws1['A2'] = "2023年度                                           单位：万元"
    ws1.merge_cells('A2:F2')

    headers = ['序号', '产品/服务类别', '2023年金额', '占比', '2022年金额', '同比增长']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, 'AI大模型平台服务', 4525.00, '36.0%', 3256.00, '39.0%'),
        (2, '智能客服解决方案', 2685.00, '21.4%', 2156.00, '24.5%'),
        (3, '数据分析服务', 1985.00, '15.8%', 1652.00, '20.2%'),
        (4, '定制化开发服务', 1568.00, '12.5%', 1256.00, '24.8%'),
        (5, '技术咨询服务', 856.00, '6.8%', 725.00, '18.1%'),
        (6, '系统集成服务', 625.00, '5.0%', 568.00, '10.0%'),
        (7, '其他业务收入', 324.50, '2.6%', 243.20, '33.4%'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [3, 5] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    total_row = 5 + len(data)
    ws1.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws1.cell(row=total_row, column=3, value=12568.50).font = HEADER_FONT
    ws1.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws1.cell(row=total_row, column=4, value='100.0%').font = HEADER_FONT
    ws1.cell(row=total_row, column=5, value=9856.20).font = HEADER_FONT
    ws1.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws1.cell(row=total_row, column=6, value='27.6%').font = HEADER_FONT
    for col in range(1, 7):
        ws1.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws1.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws1, [6, 20, 14, 10, 14, 12])

    # 按地区分类
    ws2 = wb.create_sheet("按地区分类")
    ws2['A1'] = "分地区收入统计表"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:F1')
    ws2['A2'] = "2023年度                                           单位：万元"
    ws2.merge_cells('A2:F2')

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    region_data = [
        (1, '华东地区', 5685.00, '45.2%', 4256.00, '33.6%'),
        (2, '华南地区', 2256.00, '17.9%', 1856.00, '21.6%'),
        (3, '华北地区', 1896.00, '15.1%', 1525.00, '24.3%'),
        (4, '西南地区', 1256.00, '10.0%', 1025.00, '22.5%'),
        (5, '华中地区', 856.00, '6.8%', 712.00, '20.2%'),
        (6, '其他地区', 619.50, '4.9%', 482.20, '28.5%'),
    ]

    for row_idx, row_data in enumerate(region_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [3, 5] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    total_row = 5 + len(region_data)
    ws2.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws2.cell(row=total_row, column=3, value=12568.50).font = HEADER_FONT
    ws2.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws2.cell(row=total_row, column=4, value='100.0%').font = HEADER_FONT
    ws2.cell(row=total_row, column=5, value=9856.20).font = HEADER_FONT
    ws2.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws2.cell(row=total_row, column=6, value='27.6%').font = HEADER_FONT
    for col in range(1, 7):
        ws2.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws2.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws2, [6, 20, 14, 10, 14, 12])

    wb.save(os.path.join(OUTPUT_DIR, "收入明细表.xlsx"))
    print("✓ 收入明细表.xlsx")

def create_cost_expense():
    """创建成本费用明细表"""
    wb = Workbook()

    # 成本构成
    ws1 = wb.active
    ws1.title = "成本构成"

    ws1['A1'] = "成本构成明细表"
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells('A1:E1')
    ws1['A2'] = "2023年度                                           单位：万元"
    ws1.merge_cells('A2:E2')

    headers = ['序号', '成本项目', '2023年金额', '占比', '2022年金额']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, '人工成本', 2856.80, '41.4%', 2256.50),
        (2, '云计算资源成本', 1526.50, '22.1%', 1185.20),
        (3, '数据采购成本', 856.00, '12.4%', 652.30),
        (4, '外包服务成本', 625.00, '9.1%', 525.60),
        (5, '软硬件摊销', 456.00, '6.6%', 385.20),
        (6, '技术服务费', 325.00, '4.7%', 268.50),
        (7, '其他成本', 247.00, '3.6%', 289.50),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [3, 5] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    total_row = 5 + len(data)
    ws1.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws1.cell(row=total_row, column=3, value=6892.30).font = HEADER_FONT
    ws1.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws1.cell(row=total_row, column=4, value='100.0%').font = HEADER_FONT
    ws1.cell(row=total_row, column=5, value=5562.80).font = HEADER_FONT
    ws1.cell(row=total_row, column=5).number_format = '#,##0.00'
    for col in range(1, 6):
        ws1.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws1.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws1, [6, 18, 14, 10, 14])

    # 期间费用
    ws2 = wb.create_sheet("期间费用")
    ws2['A1'] = "期间费用明细表"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:E1')
    ws2['A2'] = "2023年度                                           单位：万元"
    ws2.merge_cells('A2:E2')

    headers2 = ['序号', '费用项目', '2023年金额', '占比', '2022年金额']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    expense_data = [
        (1, '销售费用合计', 1125.50, '-', 856.20),
        ('', '  职工薪酬', 456.00, '', 356.00),
        ('', '  市场推广费', 325.00, '', 256.00),
        ('', '  差旅费', 186.00, '', 125.00),
        ('', '  业务招待费', 98.00, '', 75.00),
        ('', '  其他', 60.50, '', 44.20),
        (2, '管理费用合计', 1456.80, '-', 1125.60),
        ('', '  职工薪酬', 685.00, '', 525.00),
        ('', '  办公费', 225.00, '', 186.00),
        ('', '  折旧摊销', 186.00, '', 156.00),
        ('', '  中介服务费', 156.00, '', 98.00),
        ('', '  其他', 204.80, '', 160.60),
        (3, '财务费用合计', 186.50, '-', 125.60),
        ('', '  利息支出', 156.00, '', 105.00),
        ('', '  手续费', 18.50, '', 12.60),
        ('', '  其他', 12.00, '', 8.00),
    ]

    for row_idx, row_data in enumerate(expense_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [3, 5] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if '合计' in str(row_data[1]):
                cell.font = HEADER_FONT

    set_column_width(ws2, [6, 18, 14, 10, 14])

    wb.save(os.path.join(OUTPUT_DIR, "成本费用明细表.xlsx"))
    print("✓ 成本费用明细表.xlsx")

def create_rd_expense():
    """创建研发费用明细表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "研发费用明细"

    ws['A1'] = "研发费用明细表"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    ws['A2'] = "2023年度                                           单位：万元"
    ws.merge_cells('A2:E2')

    headers = ['序号', '研发项目', '费用金额', '占比', '费用类型']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, '大模型基础能力研发', 685.00, '36.9%', '自主研发'),
        (2, '多模态融合技术研发', 425.00, '22.9%', '自主研发'),
        (3, '行业大模型应用研发', 312.00, '16.8%', '自主研发'),
        (4, '智能客服系统研发', 186.00, '10.0%', '自主研发'),
        (5, '数据标注平台研发', 125.00, '6.7%', '自主研发'),
        (6, '安全合规技术研发', 68.50, '3.7%', '自主研发'),
        (7, '联合研发项目', 55.00, '3.0%', '合作研发'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx == 3 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    total_row = 5 + len(data)
    ws.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws.cell(row=total_row, column=3, value=1856.50).font = HEADER_FONT
    ws.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws.cell(row=total_row, column=4, value='100.0%').font = HEADER_FONT
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws, [6, 24, 14, 10, 12])

    wb.save(os.path.join(OUTPUT_DIR, "研发费用明细表.xlsx"))
    print("✓ 研发费用明细表.xlsx")

def create_fixed_assets():
    """创建固定资产明细表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "固定资产明细"

    ws['A1'] = "固定资产明细表"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    ws['A2'] = "截至2023年12月31日                                    单位：万元"
    ws.merge_cells('A2:H2')

    headers = ['序号', '资产名称', '资产类别', '原值', '累计折旧', '净值', '使用状态', '存放地点']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        (1, 'GPU服务器A100', '电子设备', 856.00, 214.00, 642.00, '在用', '杭州机房'),
        (2, 'GPU服务器H800', '电子设备', 520.00, 65.00, 455.00, '在用', '杭州机房'),
        (3, '通用服务器', '电子设备', 186.00, 56.00, 130.00, '在用', '杭州机房'),
        (4, '存储设备', '电子设备', 125.00, 31.25, 93.75, '在用', '杭州机房'),
        (5, '网络设备', '电子设备', 68.00, 25.60, 42.40, '在用', '杭州机房'),
        (6, '办公电脑', '电子设备', 85.00, 42.50, 42.50, '在用', '办公区'),
        (7, '办公家具', '办公设备', 56.00, 22.40, 33.60, '在用', '办公区'),
        (8, '空调设备', '办公设备', 35.00, 14.00, 21.00, '在用', '办公区'),
        (9, '其他设备', '电子设备', 45.00, 27.00, 18.00, '在用', '办公区'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx in [4, 5, 6] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    total_row = 5 + len(data)
    ws.cell(row=total_row, column=1, value='合计').font = HEADER_FONT
    ws.cell(row=total_row, column=4, value=1976.00).font = HEADER_FONT
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5, value=497.75).font = HEADER_FONT
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6, value=1478.25).font = HEADER_FONT
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    set_column_width(ws, [6, 18, 12, 12, 12, 12, 10, 12])

    wb.save(os.path.join(OUTPUT_DIR, "固定资产明细表.xlsx"))
    print("✓ 固定资产明细表.xlsx")

def create_three_year_comparison():
    """创建三年一期财务对比表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "三年一期对比"

    ws['A1'] = "三年一期财务指标对比表"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    ws['A2'] = "单位：万元"
    ws.merge_cells('A2:E2')

    headers = ['财务指标', '2021年', '2022年', '2023年', '2023年1-9月']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    data = [
        ('一、经营规模', '', '', '', ''),
        ('营业收入', 6525.80, 9856.20, 12568.50, 9256.80),
        ('营业成本', 3856.50, 5562.80, 6892.30, 5056.20),
        ('利润总额', 512.60, 919.64, 1288.22, 925.60),
        ('净利润', 435.71, 781.69, 1094.99, 786.76),
        ('', '', '', '', ''),
        ('二、资产规模', '', '', '', ''),
        ('资产总额', 5625.80, 8210.60, 11417.85, 10568.50),
        ('负债总额', 2856.20, 4215.80, 5225.50, 4856.20),
        ('净资产', 2769.60, 3994.80, 6192.35, 5712.30),
        ('', '', '', '', ''),
        ('三、盈利能力', '', '', '', ''),
        ('毛利率', '40.9%', '43.5%', '45.2%', '45.4%'),
        ('净利率', '6.7%', '7.9%', '8.7%', '8.5%'),
        ('ROE', '15.7%', '19.6%', '17.7%', '13.8%'),
        ('', '', '', '', ''),
        ('四、营运能力', '', '', '', ''),
        ('应收账款周转率', '4.2次', '3.8次', '3.5次', '3.6次'),
        ('存货周转率', '8.5次', '9.2次', '8.8次', '8.6次'),
        ('总资产周转率', '1.2次', '1.2次', '1.1次', '1.0次'),
        ('', '', '', '', ''),
        ('五、偿债能力', '', '', '', ''),
        ('资产负债率', '50.8%', '51.3%', '45.8%', '46.0%'),
        ('流动比率', '1.52', '1.58', '1.95', '1.88'),
        ('速动比率', '1.35', '1.42', '1.76', '1.70'),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if str(row_data[0]).startswith(('一、', '二、', '三、', '四、', '五、')):
                cell.font = HEADER_FONT
                cell.fill = ALT_FILL

    set_column_width(ws, [20, 14, 14, 14, 14])

    wb.save(os.path.join(OUTPUT_DIR, "三年一期财务对比表.xlsx"))
    print("✓ 三年一期财务对比表.xlsx")

def main():
    print("=" * 50)
    print("生成尽调财务分析 Excel 文件")
    print("=" * 50)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    create_financial_statements()
    create_receivables()
    create_payables()
    create_related_party()
    create_revenue()
    create_cost_expense()
    create_rd_expense()
    create_fixed_assets()
    create_three_year_comparison()

    print("=" * 50)
    print("完成！共生成 9 个 Excel 文件")

if __name__ == "__main__":
    main()
